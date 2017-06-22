import scrapy
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor

class indicesSpider(scrapy.Spider):
		name = "indices"
		allowed_domains = ["http://www.excelcontabilidade.com.br"]

		def start_requests(self):

			file = "M:\Indices\Indices.xlsx"
			wb = Workbook()
			wb.create_sheet(title="IGP-DI")
			wb.create_sheet(title="IGP-M")
			wb.create_sheet(title="IPCA-e")
			wb.create_sheet(title="INPC(IBGE)")
			wb.create_sheet(title="TR sem IPC's")
			wb.create_sheet(title="IPC(FIPE)")
			###wb.create_sheet(title="CDI")
			wb.create_sheet(title="TJ-Acre")
			wb.create_sheet(title="TJ-Alagoas")
			wb.create_sheet(title="TJ-Amapá")
			wb.create_sheet(title="TJ-Amazonas")
			wb.create_sheet(title="TJ-Bahia")
			wb.create_sheet(title="TJ-Ceará")
			wb.create_sheet(title="TJ-Distrito Federal")
			wb.create_sheet(title="TJ-Espirito Santo")
			wb.create_sheet(title="TJ-Goiás")
			wb.create_sheet(title="TJ-Maranhão")
			wb.create_sheet(title="TJ-Mato Grosso")
			wb.create_sheet(title="TJ-Mato Grosso do Sul")
			wb.create_sheet(title="TJ-Minas Gerais")
			wb.create_sheet(title="TJ-Pará")
			wb.create_sheet(title="TJ-Paraíba")
			wb.create_sheet(title="TJ-Paraná")
			wb.create_sheet(title="TJ-Pernambuco")
			wb.create_sheet(title="TJ-Piauí")
			wb.create_sheet(title="TJ-Rio de Janeiro")
			wb.create_sheet(title="TJ-Rio Grande do Norte")
			wb.create_sheet(title="TJ-Rio Grande do Sul")
			wb.create_sheet(title="TJ-Rondônia")
			wb.create_sheet(title="TJ-Roraima")
			wb.create_sheet(title="TJ-Santa Catarina")
			wb.create_sheet(title="TJ-São Paulo")
			wb.create_sheet(title="TJ-Sergipe")
			wb.create_sheet(title="TJ-Tocantins")
			d = wb.get_sheet_by_name("Sheet")
			wb.remove_sheet(d)
			wb.save(filename=file)

			urls = [
			"""http://www.excelcontabilidade.com.br/coeficiente/1/IGP-M""",
			"""http://www.excelcontabilidade.com.br/coeficiente/3/IGP-DI%20com%20IPIC's""",
			"""http://www.excelcontabilidade.com.br/coeficiente/4/INPC(IBGE)""",
			"""http://www.excelcontabilidade.com.br/coeficiente/6/IPC(FIPE)""",
			"""http://www.excelcontabilidade.com.br/coeficiente/8/IPCA-e""",
			"""http://www.excelcontabilidade.com.br/coeficiente/11/TR%20sem%20IPC's""",
			"""http://www.excelcontabilidade.com.br/coeficiente/14/TJ-Acre""",
			"""http://www.excelcontabilidade.com.br/coeficiente/15/TJ-Alagoas""",
			"""http://www.excelcontabilidade.com.br/coeficiente/16/TJ-Amap%C3%A1""",
			"""http://www.excelcontabilidade.com.br/coeficiente/17/TJ-Amazonas""",
			"""http://www.excelcontabilidade.com.br/coeficiente/18/TJ-Bahia""",
			"""http://www.excelcontabilidade.com.br/coeficiente/19/TJ-Cear%C3%A1""",
			"""http://www.excelcontabilidade.com.br/coeficiente/20/TJ-Distrito%20Federal""",
			"""http://www.excelcontabilidade.com.br/coeficiente/21/TJ-Espirito%20Santo""",
			"""http://www.excelcontabilidade.com.br/coeficiente/22/TJ-Goi%C3%A1s""",
			"""http://www.excelcontabilidade.com.br/coeficiente/23/TJ-Maranh%C3%A3o""",
			"""http://www.excelcontabilidade.com.br/coeficiente/24/TJ-Mato%20Grosso""",
			"""http://www.excelcontabilidade.com.br/coeficiente/25/TJ-Mato%20Grosso%20do%20Sul""",
			"""http://www.excelcontabilidade.com.br/coeficiente/26/TJ-Minas%20Gerais""",
			"""http://www.excelcontabilidade.com.br/coeficiente/27/TJ-Par%C3%A1""",
			"""http://www.excelcontabilidade.com.br/coeficiente/28/TJ-Para%C3%ADba""",
			"""http://www.excelcontabilidade.com.br/coeficiente/29/TJ-Paran%C3%A1""",
			"""http://www.excelcontabilidade.com.br/coeficiente/30/TJ-Pernambuco""",
			"""http://www.excelcontabilidade.com.br/coeficiente/31/TJ-Piau%C3%AD""",
			"""http://www.excelcontabilidade.com.br/coeficiente/32/TJ-Rio%20de%20Janeiro""",
			"""http://www.excelcontabilidade.com.br/coeficiente/42/TJ-Rio%20Grande%20do%20Norte""",
			"""http://www.excelcontabilidade.com.br/coeficiente/33/TJ-Rio%20Grande%20do%20Sul""",
			"""http://www.excelcontabilidade.com.br/coeficiente/34/TJ-Rond%C3%B4nia""",
			"""http://www.excelcontabilidade.com.br/coeficiente/35/TJ-Roraima""",
			"""http://www.excelcontabilidade.com.br/coeficiente/37/TJ-Santa%20Catarina""",
			"""http://www.excelcontabilidade.com.br/coeficiente/38/TJ-S%C3%A3o%20Paulo""",
			"""http://www.excelcontabilidade.com.br/coeficiente/39/TJ-Sergipe""",
			"""http://www.excelcontabilidade.com.br/coeficiente/40/TJ-Tocantins"""
			]


			for url in urls:
				if url == """http://www.excelcontabilidade.com.br/coeficiente/1/IGP-M""":
					yield scrapy.Request(url=url, callback=self.IGP_M)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/3/IGP-DI%20com%20IPIC's""":
					yield scrapy.Request(url=url, callback=self.IGP_DI)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/4/INPC(IBGE)""":
					yield scrapy.Request(url=url, callback=self.INPC_IBGE)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/6/IPC(FIPE)""":
					yield scrapy.Request(url=url, callback=self.IPC_FIPE)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/8/IPCA-e""":
					yield scrapy.Request(url=url, callback=self.IPCA_e)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/11/TR%20sem%20IPC's""":
					yield scrapy.Request(url=url, callback=self.TR_sem_IPC)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/14/TJ-Acre""":
					yield scrapy.Request(url=url, callback=self.TJ_ACRE)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/15/TJ-Alagoas""":
					yield scrapy.Request(url=url, callback=self.TJ_ALAGOAS)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/16/TJ-Amap%C3%A1""":
					yield scrapy.Request(url=url, callback=self.TJ_AMAPA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/17/TJ-Amazonas""":
					yield scrapy.Request(url=url, callback=self.TJ_AMAZONAS)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/18/TJ-Bahia""":
					yield scrapy.Request(url=url, callback=self.TJ_BAHIA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/19/TJ-Cear%C3%A1""":
					yield scrapy.Request(url=url, callback=self.TJ_CEARA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/20/TJ-Distrito%20Federal""":
					yield scrapy.Request(url=url, callback=self.TJ_DISTRITO_FEDERAL)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/21/TJ-Espirito%20Santo""":
					yield scrapy.Request(url=url, callback=self.TJ_ESPIRITO_SANTO)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/22/TJ-Goi%C3%A1s""":
					yield scrapy.Request(url=url, callback=self.TJ_GOIAS)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/23/TJ-Maranh%C3%A3o""":
					yield scrapy.Request(url=url, callback=self.TJ_MARANHAO)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/24/TJ-Mato%20Grosso""":
					yield scrapy.Request(url=url, callback=self.TJ_MATO_GROSSO)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/25/TJ-Mato%20Grosso%20do%20Sul""":
					yield scrapy.Request(url=url, callback=self.TJ_MATO_GROSSO_DO_SUL)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/26/TJ-Minas%20Gerais""":
					yield scrapy.Request(url=url, callback=self.TJ_MINAS_GERAIS)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/27/TJ-Par%C3%A1""":
					yield scrapy.Request(url=url, callback=self.TJ_PARA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/28/TJ-Para%C3%ADba""":
					yield scrapy.Request(url=url, callback=self.TJ_PARAIBA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/29/TJ-Paran%C3%A1""":
					yield scrapy.Request(url=url, callback=self.TJ_PARANA)
				elif url =="""http://www.excelcontabilidade.com.br/coeficiente/30/TJ-Pernambuco""":
					yield scrapy.Request(url=url, callback=self.TJ_PERNAMBUCO)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/31/TJ-Piau%C3%AD""":
					yield scrapy.Request(url=url, callback=self.TJ_PIAUI)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/32/TJ-Rio%20de%20Janeiro""":
					yield scrapy.Request(url=url, callback=self.TJ_RIO_DE_JANEIRO)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/42/TJ-Rio%20Grande%20do%20Norte""":
					yield scrapy.Request(url=url, callback=self.TJ_RIO_GRANDE_DO_NORTE)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/33/TJ-Rio%20Grande%20do%20Sul""":
					yield scrapy.Request(url=url, callback=self.TJ_RIO_GRANDE_DO_SUL)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/34/TJ-Rond%C3%B4nia""":
					yield scrapy.Request(url=url, callback=self.TJ_RONDONIA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/35/TJ-Roraima""":
					yield scrapy.Request(url=url, callback=self.TJ_RORAIMA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/37/TJ-Santa%20Catarina""":
					yield scrapy.Request(url=url, callback=self.TJ_SANTA_CATARINA)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/38/TJ-S%C3%A3o%20Paulo""":
					yield scrapy.Request(url=url, callback=self.TJ_SAO_PAULO)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/39/TJ-Sergipe""":
					yield scrapy.Request(url=url, callback=self.TJ_SERGIPE)
				elif url == """http://www.excelcontabilidade.com.br/coeficiente/40/TJ-Tocantins""":
					yield scrapy.Request(url=url, callback=self.TJ_TOCANTINS)

		def IGP_M(self, response):
			
			file = "M:\Indices\Indices.xlsx" 
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("IGP-M")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)

		def IGP_DI(self, response):
			
			file = "M:\Indices\Indices.xlsx" 
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("IGP-DI")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)

		def INPC_IBGE(self, response):
			
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("INPC(IBGE)")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)

			
		def IPC_FIPE(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("IPC(FIPE)")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)


		def IPCA_e(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("IPCA-e")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)

		def TR_sem_IPC(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TR sem IPC's")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_ACRE(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Acre")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)

		def TJ_ALAGOAS(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Alagoas")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_AMAPA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Amapá")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_AMAZONAS(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Amazonas")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_BAHIA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Bahia")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_CEARA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Ceará")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_DISTRITO_FEDERAL(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Distrito Federal")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_ESPIRITO_SANTO(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Espirito Santo")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_GOIAS(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Goiás")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_MARANHAO(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Maranhão")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass

			wb.save(filename=file)
		def TJ_MATO_GROSSO(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Mato Grosso")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_MATO_GROSSO_DO_SUL(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Mato Grosso do Sul")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_MINAS_GERAIS(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Minas Gerais")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_PARA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Pará")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_PARAIBA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Paraíba")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_PARANA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Paraná")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_PERNAMBUCO(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Pernambuco")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_PIAUI(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Piauí")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_RIO_DE_JANEIRO(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Rio de Janeiro")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_RIO_GRANDE_DO_NORTE(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Rio Grande do Norte")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_RIO_GRANDE_DO_SUL(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Rio Grande do Sul")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_RONDONIA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Rondônia")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_RORAIMA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Roraima")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_SANTA_CATARINA(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Santa Catarina")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_SAO_PAULO(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-São Paulo")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_SERGIPE(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Sergipe")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
		def TJ_TOCANTINS(self, response):
			file = "M:\Indices\Indices.xlsx"
			wb = load_workbook(file)
			ws = wb.get_sheet_by_name("TJ-Tocantins")

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws.append(cabecalho)
		
			
			dados = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						dados.append(mask)

				except(IndexError):
					break

			for line in dados:
				if line is '':
					dados.remove(line)

			for i,ano in enumerate(dados):
				try:
					if int(ano) in anos: 
						jan = dados[i+1]
						fev = dados[i+2]
						mar = dados[i+3]
						abr = dados[i+4]
						mai = dados[i+5]
						jun = dados[i+6]
						jul = dados[i+7]
						ago = dados[i+8]
						stb = dados[i+9]
						out = dados[i+10]
						nov = dados[i+11]
						dez = dados[i+12]
						meses = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws.append(meses)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)

			   
