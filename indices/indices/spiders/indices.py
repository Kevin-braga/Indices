import scrapy
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from scrapy.spiders import CrawlSpider, Rule
from scrapy.linkextractors import LinkExtractor

class indicesSpider(scrapy.Spider):
		name = "indices"
		allowed_domains = ["http://www.excelcontabilidade.com.br"]

		def start_requests(self):
		
				urls = [
				"""http://www.excelcontabilidade.com.br/coeficiente/3/IGP-DI%20com%20IPIC's""",
				"http://www.excelcontabilidade.com.br/coeficiente/4/INPC(IBGE)"
				]

				file = "indices.xlsx" 
				wb = Workbook()

				ws2 = wb.create_sheet()
				ws2.title = "IIIAISIDA"



				for url in urls:
					if url == """http://www.excelcontabilidade.com.br/coeficiente/4/INPC(IBGE)""":
						yield scrapy.Request(url=url, callback=self.IGP)
	
		def IGP(self, response):
			file = "indices.xlsx" 
			wb = Workbook()

			ws1 = wb.active
			ws1.title = "IGP"

			cabecalho = [" " , "JANEIRO", "FEVEREIRO" , "MARÇO" , "ABRIL", "MAIO" , "JUNHO" , "JULHO" , "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO" , "DEZEMBRO" ]
			ws1.append(cabecalho)
		
			
			igp = []
			anos =  list(range(1964,2018))

			jan = " "
			fev = " "
			mar = " "
			abr = " "
			mai = " "
			jun = " "
			jul = " "
			ago = " "
			stb = " "
			out = " "
			nov = " "
			dez = " "
		
			table = response.xpath("""//body//tr//text()""").extract()
			
			for i,line in enumerate(table):
				line = line.replace('\n','')
				line = line.strip()
				try:
					if line in str(anos):
						get_next=table[i+1]
						mask = (get_next.lstrip())
						mask = mask.replace('\n','')
						mask = mask.replace(' ','')								
						igp.append(mask)

				except(IndexError):
					break

			for line in igp:
				if line is '':
					igp.remove(line)

			for i,ano in enumerate(igp):
				try:
					if int(ano) in anos: 
						jan = igp[i+1]
						fev = igp[i+2]
						mar = igp[i+3]
						abr = igp[i+4]
						mai = igp[i+5]
						jun = igp[i+6]
						jul = igp[i+7]
						ago = igp[i+8]
						stb = igp[i+9]
						out = igp[i+10]
						nov = igp[i+11]
						dez = igp[i+12]
						dados = [ano,jan,fev,mar,abr,mai,jun,jul,ago,stb,out,nov,dez]

						for row in range(1):
							ws1.append(dados)
					else:
						pass
				except ValueError:
					pass
			
			wb.save(filename=file)
			print("IGP COMPLETO")

		
			
		# def IGP-DI(self, response):

		# def TR(self, response):

		# def IGP-M(self, response):

		# def IPCA-E(self, response):

		# def IPC FIPE(self, response):

		# def CDI(self, response):


			   
